#!/usr/bin/env python3
"""
인스턴스 리소스 사용률 분석 스크립트  v2
사용법:
  python analyze.py [데이터폴더] [사양파일.xlsx] [옵션]
  python analyze.py ./data 인망_행망_인스턴스자원현황2.xlsx
  python analyze.py ./data --cpu-target 65 --mem-target 75
옵션:
  --cpu-target  CPU 목표 사용률 % (기본 70)
  --mem-target  MEM 목표 사용률 % (기본 80)
  --out         출력 파일명 (기본: 리소스_분석보고서_YYYYMMDD_HHMM.xlsx)
"""
import sys, os, argparse, warnings, unicodedata
from pathlib import Path
from datetime import datetime

warnings.filterwarnings('ignore')
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install pandas openpyxl --break-system-packages -q")
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

NHN_CPU = [2, 4, 8, 16, 32, 64]

# NHN Cloud 인스턴스 사양 테이블 (vCPU, RAM_GB)
# 이미지 기준: m2(1:2), c2(1:1), r2(1:4~8), x1(고메모리)
NHN_SPECS = [
    # m2 계열 (vCPU:RAM = 1:2)
    (2,   4),
    (4,   8),
    (8,  16),
    (16, 32),
    (32, 64),
    # c2 계열 (vCPU:RAM = 1:1)
    (2,  2),
    (4,  4),
    (8,  8),
    (16,16),
    # r2 계열 (vCPU:RAM = 1:4~8)
    (2,  8),
    (4, 16),
    (8, 32),
    (8, 64),
    # x1 계열 (고메모리)
    (16, 64),
    (16,128),
    (32,128),
    (32,256),
    (64,256),
]

def find_best_spec(need_cpu, need_ram):
    """
    CPU와 RAM 요구량을 모두 충족하는 NHN Cloud 사양 중
    비용(vCPU 기준) 최소 사양 반환. 동일 vCPU면 RAM 작은 것 우선.
    """
    candidates = [
        (c, r) for (c, r) in NHN_SPECS
        if c >= need_cpu and r >= need_ram
    ]
    if not candidates:
        # 요구량 초과 시 가장 큰 사양
        return max(NHN_SPECS, key=lambda x: (x[0], x[1]))
    # vCPU 최소 → 동일 vCPU면 RAM 최소
    return min(candidates, key=lambda x: (x[0], x[1]))

# ── 파일 분류 (NFC 정규화 필수)
def nfc(s): return unicodedata.normalize('NFC', str(s))

def classify_file(name):
    n = nfc(name).lower()
    net   = '인망' if '인망' in n else '행망' if '행망' in n else None
    ftype = 'cpu'  if 'cpu_mem' in n else 'net' if 'network' in n else None
    stat  = 'max'  if '최대' in n else 'avg' if '평균' in n else None
    return (net, ftype, stat) if net and ftype and stat else None

def is_spec_file(name):
    n = nfc(name).lower()
    return any(k in n for k in ['사양','spec','자원현황','인스턴스현황'])

def is_pod_file(path):
    """xlsx 파일을 열어 podName 컬럼이 있으면 Pod 파일로 판단"""
    try:
        df = pd.read_excel(path, nrows=2, engine='openpyxl')
        return 'podName' in df.columns
    except:
        return False

# agentPcode → 클러스터명 매핑
PCODE_CLUSTER = {
    22: '인망 DMZ Cluster',
    24: '행망 AP Cluster',
}

def load_pod_files(fpaths):
    """Pod 파일 목록을 로드하고 RS별 집계"""
    import re as _re
    from datetime import date as _date

    def _dates_from_fname(fname):
        """파일명에서 날짜 추출 (YYYYMMDD 또는 YYMMDD, 오타 대응)"""
        m = _re.match(r'^(\d{6,8})-(\d{6,9})_', nfc(fname))
        if not m: return []
        result = []
        for d in [m.group(1)[:8], m.group(2)[:8]]:
            try:
                if len(d) == 8:
                    result.append(_date(int(d[:4]), int(d[4:6]), int(d[6:8])))
                elif len(d) == 6:
                    result.append(_date(2000+int(d[:2]), int(d[2:4]), int(d[4:6])))
            except: pass
        return result

    # 수치형으로 변환할 컬럼 목록
    NUM_COLS = [
        'cpu_request', 'cpu_total_milli', 'mem_working_set',
        'cpu_per_request', 'cpu_per_limit',
        'memory_request', 'memory_limit',
        'memory_per_request', 'memory_per_limit',
    ]
    dfs = []
    _pod_dates = set()

    for p in fpaths:
        try:
            df = pd.read_excel(str(p), engine='openpyxl')
            if 'podName' not in df.columns:
                continue
            df['time'] = pd.to_datetime(df['time'], errors='coerce').dt.date
            for c in NUM_COLS:
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors='coerce')
            # 망 구분
            fname = nfc(str(p).split('/')[-1])
            fname_l = fname.lower()
            if '인망' in fname_l:
                df['netlabel'] = '인망'
            elif '행망' in fname_l:
                df['netlabel'] = '행망'
            else:
                df['netlabel'] = df['agentPcode'].map({22:'인망',24:'행망'}).fillna('미분류')
            df['cluster'] = df['agentPcode'].map(PCODE_CLUSTER).fillna('Unknown Cluster')
            # Deployment / DaemonSet / replicaSetName 없으면 빈 문자열
            for col in ['Deployment', 'DaemonSet', 'replicaSetName']:
                if col not in df.columns:
                    df[col] = ''
            dfs.append(df)
            # 파일명에서 날짜 추출
            for d in _dates_from_fname(fname):
                _pod_dates.add(d)
        except Exception as e:
            print(f"  ❌ Pod 파일 로드 실패: {p} — {e}")

    if not dfs:
        return pd.DataFrame(), []
    raw = pd.concat(dfs, ignore_index=True)
    raw = raw.dropna(subset=['podName'])
    _pod_dates_sorted = sorted(_pod_dates)


    # replicaSetName 없으면 podName을 그룹키로 사용 (DaemonSet 등)
    raw['_grp_key'] = raw.apply(
        lambda r: str(r['replicaSetName']).strip()
                  if str(r.get('replicaSetName','')).strip() not in ('','nan','None')
                  else str(r['podName']),
        axis=1
    )

    # 집계: RS(또는 Pod)별 CPU/MEM 최대
    def safe_agg(col, func):
        return (col, func) if col in raw.columns else ('podName', 'count')

    agg_dict = {}
    # Pod명 보존 (대표값 first)
    agg_dict['podName'] = ('podName', 'first')
    # 전체 Pod 수
    agg_dict['pod_count'] = ('podName', 'nunique')
    # 실사용량 → 최대
    for c in ['cpu_total_milli', 'mem_working_set']:
        if c in raw.columns: agg_dict[c+'_max'] = (c, 'max')
    # % 지표 → 최대
    for c in ['cpu_per_request', 'cpu_per_limit', 'memory_per_request', 'memory_per_limit']:
        if c in raw.columns: agg_dict[c+'_max'] = (c, 'max')
    # Request / Limit 절대값 → 최대 (설정값이므로 max=last)
    for c in ['cpu_request', 'memory_request', 'memory_limit']:
        if c in raw.columns: agg_dict[c+'_val'] = (c, 'max')
    # 워크로드 정보 → last
    for c in ['Deployment', 'DaemonSet', 'replicaSetName', 'cluster', 'netlabel', 'onodeName']:
        if c in raw.columns: agg_dict[c] = (c, 'last')

    # 활성 Pod 수: cpu_request > 0 OR cpu_total_milli > 0 인 Pod 수
    # → 구 RS 판별에 활용 (활성 비율이 너무 낮으면 구 RS로 분류)
    if 'cpu_total_milli' in raw.columns and 'cpu_request' in raw.columns:
        raw['_is_active_pod'] = (
            (raw['cpu_total_milli'].fillna(0) > 0) |
            (raw['cpu_request'].fillna(0) > 0)
        )
        agg_dict['active_pod_count'] = ('_is_active_pod', 'sum')
    else:
        raw['_is_active_pod'] = True
        agg_dict['active_pod_count'] = ('_is_active_pod', 'sum')

    grp = raw.groupby(['_grp_key','namespace']).agg(**agg_dict).reset_index()
    grp = grp.rename(columns={'_grp_key': 'rs_or_pod'})

    # MEM 단위 변환: bytes → MB
    for col in ['mem_working_set_max']:
        if col in grp.columns:
            grp[col] = (grp[col] / 1024 / 1024).round(1)
    for col in ['memory_request_val', 'memory_limit_val']:
        if col in grp.columns:
            grp[col] = (grp[col] / 1024 / 1024).round(1)  # MB

    # 수치 반올림
    for col in grp.select_dtypes('float').columns:
        grp[col] = grp[col].round(1)

    grp = grp.sort_values(['netlabel','cluster','namespace','rs_or_pod']).reset_index(drop=True)
    # RS 값 유효성 확인
    _rs_empty = (grp["rs_or_pod"] == grp["podName"]).sum() if "podName" in grp.columns else 0
    if _rs_empty == len(grp):
        print("  ⚠ replicaSetName 값이 비어있어 Pod명 기준으로 집계됩니다.")
        print("    → MXQL 쿼리에 replicaSetName 포함 후 재추출 필요")
    print(f"  ✓ Pod 데이터: {len(grp)}개 RS/Pod, {grp['namespace'].nunique()}개 namespace")
    return grp, _pod_dates_sorted

# ── 데이터 로드
def load_xlsx(path, value_cols):
    df = pd.read_excel(path, engine='openpyxl')
    needed = ['oname', 'time'] + value_cols
    df = df[[c for c in needed if c in df.columns]]
    df = df.dropna(subset=['oname'])
    df['time'] = pd.to_datetime(df['time'], errors='coerce').dt.date
    df = df[df['time'].notna()]
    for c in value_cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors='coerce')
    return df.dropna(subset=[c for c in value_cols if c in df.columns])

# ── 사양 파일 로드
def load_spec(path):
    if not path or not Path(path).exists(): return {}
    df = pd.read_excel(path, header=0, engine='openpyxl')
    df.columns = [nfc(c) for c in df.columns]
    # 컬럼 자동 감지
    def find(cands): return next((c for c in df.columns if any(k.lower() in c.lower() for k in cands)), None)
    c_inst = find(['인스턴스명','instance','hostname','oname','서버'])
    c_sys  = find(['시스템명','system','서비스'])
    c_cpu  = find(['CPU','cpu','vCPU','코어'])
    c_ram  = find(['RAM','ram','MEM','mem','GB','메모리'])
    c_net  = find(['망','네트워크','net'])
    if not c_inst: print("  ⚠ 인스턴스명 컬럼을 찾지 못했습니다."); return {}
    spec = {}
    for _, row in df.iterrows():
        inst = nfc(str(row[c_inst] or '')).strip().replace('.novalocal','').lower()
        if not inst or inst in ('nan','none'): continue
        sysn  = nfc(str(row[c_sys] or '')).strip() if c_sys else ''
        cpu_v = int(float(row[c_cpu] or 0)) if c_cpu else 0
        ram_v = int(float(row[c_ram] or 0)) if c_ram else 0
        net_v = nfc(str(row[c_net] or '')).strip() if c_net else ''
        nl = '인망' if ('인터넷' in net_v or '인망' in net_v) else \
             '행망' if ('행정' in net_v or '행망' in net_v) else ''
        entry = {'시스템명': sysn or None, 'cpu': cpu_v or None, 'ram': ram_v or None}
        if nl: spec[f'{nl}|{inst}'] = entry
        if inst not in spec: spec[inst] = entry  # 망 없는 쪽은 먼저 온 것 유지
    cnt = len([k for k in spec if '|' in k])
    print(f"  ✓ 사양 파일 로드: {cnt}대")
    return spec

def lookup_spec(spec, netlabel, oname):
    k1 = f'{netlabel}|{nfc(oname).replace(".novalocal","").lower().strip()}'
    k2 = nfc(oname).replace('.novalocal','').lower().strip()
    return spec.get(k1) or spec.get(k2)

# ── 판정 규칙
DEFAULT_RULES = [
    ('tibero',       'except', '물리장비 - 예외처리'),
    ('waf',          'except', 'WAF 보안장비'),
    ('firewall',     'except', '방화벽'),
    ('-ips-',        'except', 'IPS 보안장비'),
    ('loadrunner',   'except', '성능테스트 도구'),
    ('loadgenerator','except', '성능테스트 도구'),
    ('jmeter',       'except', '성능테스트 도구'),
    ('ngrinder',     'except', '성능테스트 도구'),
    ('mock-',        'except', 'Mock 서버'),
    ('stg-',         'except', '검증환경'),
    ('dev-',         'except', '개발환경'),
    ('nginx-test',   'except', '테스트 서버'),
    ('redis-session','except', '신규 Redis - 수집 미흡'),
    ('redis-vector', 'except', '신규 Redis - 수집 미흡'),
    ('ai-search',    'except', '신규 서비스(3/4 오픈)'),
    ('apig-imc',     'except', 'API GW - MEM 고사용 특성'),
    ('appiron',      'except', '신규 서버 - 수집 미흡'),
    ('noti-web',     'down',   '점검 페이지 - 분기별 전용'),
    ('sorry-web',    'down',   '점검 페이지 - 분기별 전용'),
    ('cicd',         'keep',   'CI/CD - 빌드 시 순간 부하 특성'),
    ('ionengs',      'keep',   '접근제어 솔루션'),
    ('static',       'keep',   'CDN 대체 정적 서버'),
    ('well-tracer',  'keep',   '사용자제어 솔루션'),
    ('int-tracer',   'keep',   '사용자제어 솔루션'),
    ('edms',         'keep',   '주요 서버 - EDMS'),
    ('whatap',       'keep',   '모니터링 솔루션'),
    ('ahnlab',       'keep',   '보안 솔루션'),
    ('syslog',       'keep',   'Syslog 서버'),
    ('intraplus',    'keep',   '관리포털 - 역할 유지'),
    ('pdf',          'keep',   '주요 서버 - PDF'),
]

def grade_row(row, rules, th):
    o = nfc(row['oname']).lower()
    for kw, action, reason in rules:
        if kw in o:
            return {'except':'예외처리','keep':'유지(확정)','down':'강력권고','review':'일시적피크→검토'}.get(action,'보통'), reason
    if row['cpu_max'] >= th['keep'] or row['mem_max'] >= th['keep']:
        return '유지(주의)', f"CPU 최대 {row['cpu_max']:.1f}% 또는 MEM 최대 {row['mem_max']:.1f}%"
    if row['cpu_max'] < th['s_max'] and row['mem_max'] < th['s_max']:  # 40% 이하 → 감설
        return '강력권고', f"CPU 최대 {row['cpu_max']:.1f}% / 평균 {row['cpu_avg']:.1f}%"
    if row['cpu_max'] < th['r_max'] and row['mem_max'] < th['r_max']:
        return '검토', f"CPU 최대 {row['cpu_max']:.1f}% / 평균 {row['cpu_avg']:.1f}%"
    return '보통', f"CPU 최대 {row['cpu_max']:.1f}%"

def snap_up(v):
    for s in NHN_CPU:
        if s >= v: return s
    return NHN_CPU[-1]

def snap_down(v):
    for s in reversed(NHN_CPU):
        if s < v: return s
    return NHN_CPU[0]

def calc_rec(row, tc, tm):
    """
    스케일 권고 계산.
    CPU와 RAM을 독립적으로 필요량 산출 후 NHN Cloud 사양 테이블에서 최적 사양 탐색.
    - 유지(주의): 70% 이상 증설 검토 대상 -> 스케일 업 계산
    - 다운 후 예상 CPU > tc% 또는 예상 MEM > tm% 이면 현상태유지(다운불가)
    - CPU만 낮춰도 되는 경우, MEM만 낮춰도 되는 경우 모두 처리
    """
    pj = row['판정']
    if pj in ('예외처리', '유지(확정)', '보통'):
        return None, None, pj, None, None
    cur_c = row.get('현재_cpu'); cur_r = row.get('현재_ram')
    if not cur_c or not cur_r or str(cur_c)=='nan': return None, None, '확인필요', None, None
    cur_c = int(float(cur_c)); cur_r = int(float(cur_r))

    # CPU/MEM 각각 목표 달성을 위한 최소 필요량 계산
    need_cpu = (cur_c * row['cpu_max']) / tc   # CPU 목표 달성 최소 vCPU
    need_ram = (cur_r * row['mem_max']) / tm   # MEM 목표 달성 최소 RAM(GB)

    if pj in ('강력권고', '검토', '일시적피크→검토'):
        # ── 다운 방향: 필요량 충족하는 최소 사양 탐색
        rec_c, rec_r = find_best_spec(need_cpu, need_ram)

        # 현재보다 커지거나 같으면 다운 불가
        if rec_c >= cur_c and rec_r >= cur_r:
            # 사양 테이블에서 현재보다 한 단계 아래 탐색
            lower = [(c,r) for (c,r) in NHN_SPECS if c <= cur_c and r <= cur_r and (c < cur_c or r < cur_r)]
            if not lower:
                return None, None, '현상태유지', None, None
            rec_c, rec_r = max(lower, key=lambda x: (x[0], x[1]))

        # 예상 사용률 검증
        est_cpu = round(row['cpu_max'] * cur_c / rec_c, 1) if rec_c > 0 else 999
        est_mem = round(row['mem_max'] * cur_r / rec_r, 1) if rec_r > 0 else 999

        if est_cpu > tc or est_mem > tm:
            return None, None, '현상태유지(다운불가)', est_cpu, est_mem

        # 스케일 방향 세분화
        if rec_c < cur_c and rec_r < cur_r:
            direction = '스케일 다운'
        elif rec_c < cur_c:
            direction = '스케일 다운 (CPU)'
        elif rec_r < cur_r:
            direction = '스케일 다운 (MEM)'
        else:
            return None, None, '현상태유지', None, None

        return rec_c, rec_r, direction, est_cpu, est_mem

    else:
        # 유지(주의): 70% 이상 증설 검토 대상 -> 스케일 업 계산
        rec_c, rec_r = find_best_spec(need_cpu, need_ram)
        # 현재보다 커야 의미 있음
        if rec_c <= cur_c and rec_r <= cur_r:
            upper = [(c,r) for (c,r) in NHN_SPECS if c > cur_c or r > cur_r]
            if not upper:
                return None, None, '현상태유지', None, None
            rec_c, rec_r = min(upper, key=lambda x: (x[0], x[1]))
        est_cpu = round(row['cpu_max'] * cur_c / rec_c, 1) if rec_c > 0 else 0
        est_mem = round(row['mem_max'] * cur_r / rec_r, 1) if rec_r > 0 else 0
        return rec_c, rec_r, '스케일 업', est_cpu, est_mem

# ── 스타일
C = dict(
    NAVY='1F3864', BLUE='2E4D8C',
    GRN='C8E6C9', GRN_T='166534', YEL='FFD54F', YEL_T='854D0E',
    RED='FF8A80', RED_T='991B1B', TEAL='B2EBF2', TEAL_T='155E75',
    ORG='FFE0B2', ORG_T='9A3412', LGRAY='F5F5F5', WHITE='FFFFFF',
    BLACK='1A1F33', SGRAY='5A6480',
    IM_BG='DBEAFE', IM_T='1E40AF', HM_BG='DCFCE7', HM_T='166534',
    DN_BG='E3F2FD', DN_T='1565C0', UP_BG='FCE4EC', UP_T='B71C1C',
    WARN='FFD0D0', WARN2='FFF8D0',
)
PJ_FILL = {
    '강력권고':(C['GRN'],C['GRN_T']), '검토':(C['YEL'],C['YEL_T']),
    '유지(확정)':(C['TEAL'],C['TEAL_T']), '유지(주의)':(C['RED'],C['RED_T']),
    '일시적피크→검토':(C['ORG'],C['ORG_T']),
    '예외처리':(C['LGRAY'],C['SGRAY']), '보통':(C['WHITE'],C['BLACK']),
}
SCALE_FILL = {
    '스케일 다운':(C['DN_BG'],C['DN_T']),
    '스케일 업':  (C['UP_BG'],C['UP_T']),
}
thin = Side(style='thin', color='CCCCCC')
def bdr(): return Border(left=thin,right=thin,top=thin,bottom=thin)
def fl(c): return PatternFill('solid',start_color=c,end_color=c)

def sc(ws,r,c,val,sz=10,bold=False,color='1A1F33',bg=None,
       ha='center',wrap=False,fmt=None,border=True):
    addr = get_column_letter(c)+str(r)
    ws[addr] = val
    ws[addr].font = Font(name='Arial',sz=sz,bold=bold,color=color)
    if bg: ws[addr].fill = fl(bg)
    ws[addr].alignment = Alignment(horizontal=ha,vertical='center',wrap_text=wrap)
    if border: ws[addr].border = bdr()
    if fmt: ws[addr].number_format = fmt

# 스타일 추가
HOLD_BG = 'FFF8E1'; HOLD_T = 'E65100'
GROUP_NOW = 'D6E4F7'; GROUP_REC = 'D5F0D5'

def bdr_med_right():
    med = Side(style='medium', color='AAAAAA')
    return Border(left=thin, right=med, top=thin, bottom=thin)

def make_detail_sheet(wb, df_src, sheet_name, tab_color, mode='full', tc=70, tm=80):
    """
    mode='full'   : 전체 상세(상세) — 모든 컬럼
    mode='simple' : 전체 상세(간소) — 핵심 컬럼만
    정렬: 인망→행망, 각 망 내 oname 알파벳 순
    """
    import pandas as _pd
    ws = wb.create_sheet(sheet_name); ws.tab_color = tab_color
    ws.sheet_view.showGridLines = False

    # ── 컬럼 정의 (헤더명, 너비, 그룹)
    fixed = [('망',5.5,None),('서버명',36,None),('시스템명',38,None),('서버유형',9.5,None)]
    spec_now = [('현재\nvCPU',6.5,'now'),('현재\nRAM(GB)',9.5,'now')]
    if mode == 'full':
        usage_now = [('CPU\n최대(%)',8,'now'),('CPU\n평균(%)',8,'now'),
                     ('MEM\n최대(%)',8,'now'),('MEM\n평균(%)',8,'now'),
                     ('CPU≥80%\n발생 날짜',30.5,'now'),('MEM≥80%\n발생 날짜',30.5,'now')]
        rec_cols  = [('판정',16,'rec'),('스케일\n방향',19,'rec'),
                     ('권고\nvCPU',6.5,'rec'),('권고\nRAM(GB)',9.5,'rec'),
                     ('스케일 후\n예상CPU(%)',11.5,'rec'),('스케일 후\n예상MEM(%)',12,'rec'),
                     ('권고 근거',40,'rec')]
    else:
        usage_now = [('CPU\n최대(%)',8,'now'),('MEM\n최대(%)',8,'now')]
        rec_cols  = [('권고\nvCPU',6.5,'rec'),('권고\nRAM(GB)',9.5,'rec'),
                     ('예상\nCPU(%)',8,'rec'),('예상\nMEM(%)',8.5,'rec')]

    all_cols = fixed + spec_now + usage_now + rec_cols
    nC = len(all_cols)
    for ci,(h,w,_) in enumerate(all_cols,1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 제목 행
    ws.merge_cells(f'A1:{get_column_letter(nC)}1')
    ws['A1'] = sheet_name
    ws['A1'].font = Font(name='Arial',sz=13,bold=True,color=C['NAVY'])
    ws['A1'].alignment = Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A2:{get_column_letter(nC)}2')
    ws['A2'] = f'목표: CPU {tc}% / MEM {tm}% | 최대값 기준 | 최한시·주말 제외 | NHN Cloud vCPU: 2→4→8→16→32→64'
    ws['A2'].font = Font(name='Arial',sz=9,color='888888')
    ws['A2'].alignment = Alignment(horizontal='left',vertical='center')
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6

    # 그룹 헤더 행 (row 4)
    ws.row_dimensions[4].height = 20
    fixed_end = len(fixed)
    now_start = fixed_end + 1
    now_end   = fixed_end + len(spec_now) + len(usage_now)
    rec_start = now_end + 1
    rec_end   = nC

    for c in range(1, fixed_end+1):
        sc(ws,4,c,'',bg=C['WHITE'],border=False)
    # "현재" 병합
    ws.merge_cells(f'{get_column_letter(now_start)}4:{get_column_letter(now_end)}4')
    addr_now = get_column_letter(now_start)+'4'
    ws[addr_now] = '현  재'
    ws[addr_now].font = Font(name='Arial',sz=10,bold=True,color='1E3A5F')
    ws[addr_now].fill = PatternFill('solid',start_color=GROUP_NOW,end_color=GROUP_NOW)
    ws[addr_now].alignment = Alignment(horizontal='center',vertical='center')
    med = Side(style='medium',color='AAAAAA')
    ws[addr_now].border = Border(left=med,right=med,top=med,bottom=thin)
    # "권고" 병합
    ws.merge_cells(f'{get_column_letter(rec_start)}4:{get_column_letter(rec_end)}4')
    addr_rec = get_column_letter(rec_start)+'4'
    ws[addr_rec] = '권  고'
    ws[addr_rec].font = Font(name='Arial',sz=10,bold=True,color='14532D')
    ws[addr_rec].fill = PatternFill('solid',start_color=GROUP_REC,end_color=GROUP_REC)
    ws[addr_rec].alignment = Alignment(horizontal='center',vertical='center')
    ws[addr_rec].border = Border(left=med,right=med,top=med,bottom=thin)

    # 컬럼 헤더 행 (row 5)
    ws.row_dimensions[5].height = 36
    for ci,(h,w,grp) in enumerate(all_cols,1):
        bg_hdr = '4472C4' if grp=='now' else ('375623' if grp=='rec' else C['NAVY'])
        sc(ws,5,ci,h,10,True,'FFFFFF',bg_hdr,'center',True)

    # 정렬: 인망→행망, 각 hostname 알파벳
    df_s = df_src.copy()
    df_s['_no'] = df_s['netlabel'].map({'인망':0,'행망':1}).fillna(2)
    df_s = df_s.sort_values(['_no','oname']).drop(columns='_no').reset_index(drop=True)

    for ri,row in df_s.iterrows():
        R = ri+6; ws.row_dimensions[R].height = 18
        bg = C['LGRAY'] if ri%2==0 else C['WHITE']
        nl = row.get('netlabel','')
        nb = C['IM_BG'] if nl=='인망' else C['HM_BG']
        nt = C['IM_T']  if nl=='인망' else C['HM_T']
        def cpub(v): return C['WARN'] if v>=80 else C['WARN2'] if v>=60 else bg
        def cput(v): return C['RED_T'] if v>=80 else C['YEL_T'] if v>=60 else C['BLACK']
        pj  = str(row.get('판정',''))
        sd  = str(row.get('스케일방향','')) if _pd.notna(row.get('스케일방향')) else '-'
        rv  = row.get('권고_cpu');  rv  = int(float(rv)) if _pd.notna(rv) else None
        rm  = row.get('권고_ram');  rm  = int(float(rm)) if _pd.notna(rm) else None
        ec  = row.get('예상_cpu');  ec  = float(ec) if _pd.notna(ec) else None
        em  = row.get('예상_mem');  em  = float(em) if _pd.notna(em) else None
        cm  = float(row['cpu_max']); ca2 = float(row['cpu_avg'])
        mm  = float(row['mem_max']); ma2 = float(row['mem_avg'])
        sysn= str(row.get('시스템명','')) if _pd.notna(row.get('시스템명')) else ''
        svr = str(row.get('서버유형',''))
        cur_c=row.get('현재_cpu'); cur_c=int(float(cur_c)) if _pd.notna(cur_c) else None
        cur_r=row.get('현재_ram'); cur_r=int(float(cur_r)) if _pd.notna(cur_r) else None
        if   sd=='스케일 다운':        rv_bg,rv_tc = C['DN_BG'],C['DN_T']
        elif '현상태' in sd:            rv_bg,rv_tc = HOLD_BG,HOLD_T
        elif sd=='스케일 업':          rv_bg,rv_tc = C['UP_BG'],C['UP_T']
        else:                           rv_bg,rv_tc = bg,C['SGRAY']

        col_vals = []
        # 고정
        col_vals += [(nl,nb,nt,True,None,'center'),(row['oname'],bg,C['BLACK'],False,None,'left'),
                     (sysn,bg,C['BLACK'] if sysn else C['SGRAY'],False,None,'left'),
                     (svr,bg,C['SGRAY'],False,None,'center')]
        # 현재 스펙
        col_vals += [(cur_c or '?',bg,C['BLACK'],True,None,'center'),
                     (cur_r or '?',bg,C['BLACK'],False,None,'center')]
        # 현재 사용률
        col_vals.append((cm,cpub(cm),cput(cm),cm>=60,'0.0','center'))
        if mode=='full': col_vals.append((ca2,bg,C['SGRAY'],False,'0.0','center'))
        col_vals.append((mm,cpub(mm),cput(mm),mm>=60,'0.0','center'))
        if mode=='full':
            col_vals.append((ma2,bg,C['SGRAY'],False,'0.0','center'))
            # 80% 발생 날짜
            cal = str(row.get('cpu_alert','-')) if _pd.notna(row.get('cpu_alert')) else '-'
            mal = str(row.get('mem_alert','-')) if _pd.notna(row.get('mem_alert')) else '-'
            col_vals.append((cal, C['WARN'] if cal!='-' else bg, C['RED_T'] if cal!='-' else C['SGRAY'], False, None, 'left'))
            col_vals.append((mal, C['WARN'] if mal!='-' else bg, C['RED_T'] if mal!='-' else C['SGRAY'], False, None, 'left'))
        # 권고
        if mode=='full':
            # 판정
            pj_bg,pj_tc = PJ_FILL.get(pj,(bg,C['BLACK']))
            col_vals.append((pj,pj_bg,pj_tc,True,None,'center'))
            # 스케일 방향
            if   sd=='스케일 다운':  s_bg,s_tc = C['DN_BG'],C['DN_T']
            elif '현상태' in sd:     s_bg,s_tc = HOLD_BG,HOLD_T
            elif sd=='스케일 업':    s_bg,s_tc = C['UP_BG'],C['UP_T']
            else:                    s_bg,s_tc = bg,C['SGRAY']
            col_vals.append((sd,s_bg,s_tc,True,None,'center'))
        col_vals.append((rv or '-',rv_bg,rv_tc,bool(rv),None,'center'))
        col_vals.append((rm or '-',rv_bg,rv_tc,bool(rm),None,'center'))
        col_vals.append((ec or '-',rv_bg,rv_tc,bool(ec),'0.0' if ec else None,'center'))
        col_vals.append((em or '-',rv_bg,rv_tc,bool(em),'0.0' if em else None,'center'))
        if mode=='full':
            reason = str(row.get('reason','')) if _pd.notna(row.get('reason')) else ''
            col_vals.append((reason,bg,C['SGRAY'],False,None,'left'))

        for ci,(v,b,tc2,bold,fmt,ha) in enumerate(col_vals,1):
            sc(ws,R,ci,v,9 if ci==2 else 10,bold,tc2,b,ha,ci in(2,3),fmt)

    ws.freeze_panes = 'A6'

# 하위 호환 래퍼 (server.py 등에서 make_sheet 호출 시)
def make_sheet(wb, df, title, has_spec, tc, tm, tab='AAAAAA'):
    make_detail_sheet(wb, df, title, tab, mode='full', tc=tc, tm=tm)


def make_pod_sheet(wb, pod_df, date_range=''):
    """
    Pod 분석 시트 생성.
    컬럼: 망 | 클러스터 | 네임스페이스 | Pod명 | 배포 워커노드 | Deployment | ReplicaSet |
          CPU Request(m) | CPU 실사용 최대(m) | CPU Req 대비(%) | CPU Limit 대비(%) |
          MEM Request(MB) | MEM Limit(MB) | MEM 실사용 최대(MB) | MEM Req 대비(%) | MEM Limit 대비(%) |
          비고
    """
    ws = wb.create_sheet('Pod 분석')
    ws.tab_color = 'FF9800'
    ws.sheet_view.showGridLines = False

    POD_HDRS = [
        ('망',              7 ),
        ('클러스터',        22),
        ('네임스페이스',    22),
        ('Pod명',           44),
        ('배포 워커노드',   36),
        ('Deployment',      30),
        ('ReplicaSet',      30),
        ('CPU\nRequest(m)', 11),
        ('CPU 실사용\n최대(m)', 11),
        ('CPU\nReq 대비(%)', 11),
        ('CPU\nLimit 대비(%)', 11),
        ('MEM\nRequest(MB)', 12),
        ('MEM\nLimit(MB)',   12),
        ('MEM 실사용\n최대(MB)', 12),
        ('MEM\nReq 대비(%)', 11),
        ('MEM\nLimit 대비(%)', 11),
        ('비고', 30),
    ]
    nC = len(POD_HDRS)
    for ci,(h,w) in enumerate(POD_HDRS,1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.merge_cells(f'A1:{get_column_letter(nC)}1')
    sc(ws,1,1,'Pod 리소스 분석  —  2026-01-19 / 2026-03-04 기준',
       13,True,'FFFFFF',C['NAVY'],'left',border=False)
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{get_column_letter(nC)}2')
    sc(ws,2,1,
       '※ KST 06시 이후·평일 기준 | CPU/MEM 실사용 최대(m/MB) | Request·Limit 대비 사용률(%) | 워커노드 64vCPU/256GB — 노드 단위 사용률로 Pod 부하 파악 불가',
       9,False,'888888',C['NAVY'],'left',border=False)
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6

    ws.row_dimensions[4].height = 38
    for ci,(h,w) in enumerate(POD_HDRS,1):
        sc(ws,4,ci,h,10,True,'FFFFFF',C['NAVY'],'center',True)
    ws.freeze_panes = 'A5'

    if pod_df.empty:
        ws.merge_cells(f'A5:{get_column_letter(nC)}5')
        sc(ws,5,1,'Pod 데이터 없음 — 파일 업로드 후 재분석 필요',9,False,'888888',C['LGRAY'],'center')
        ws.row_dimensions[5].height = 24
        return

    def fv(row, key, default=0.0):
        v = row.get(key, default)
        try:
            f = float(v)
            return 0.0 if pd.isna(f) else f
        except:
            return default

    def pct_bg(v, warn=80, caution=40):  # Pod 기준: 40%이하 감설 / 80%이상 증설
        return C['WARN'] if v>=warn else C['WARN2'] if v>=caution else None
    def pct_tc(v, warn=80, caution=40):
        return C['RED_T'] if v>=warn else C['YEL_T'] if v>=caution else C['BLACK']

    for ri, row in pod_df.iterrows():
        R = ri + 5
        ws.row_dimensions[R].height = 18
        bg = C['LGRAY'] if ri%2==0 else C['WHITE']
        nl = str(row.get('netlabel',''))
        nb = C['IM_BG'] if nl=='인망' else C['HM_BG']
        nt = C['IM_T']  if nl=='인망' else C['HM_T']

        cpu_req     = fv(row, 'cpu_request_val')
        cpu_used    = fv(row, 'cpu_total_milli_max')
        cpu_pct_req = fv(row, 'cpu_per_request_max')
        cpu_pct_lim = fv(row, 'cpu_per_limit_max')
        mem_req_mb  = fv(row, 'memory_request_val')
        mem_lim_mb  = fv(row, 'memory_limit_val')
        mem_used_mb = fv(row, 'mem_working_set_max')
        mem_pct_req = fv(row, 'memory_per_request_max')
        mem_pct_lim = fv(row, 'memory_per_limit_max')

        notes = []
        if cpu_pct_lim >= 80:   notes.append('⚠ CPU Limit 위험')
        elif cpu_pct_req >= 40: notes.append('△ CPU Request 40% 초과')  # Pod 감설 기준
        if mem_pct_lim >= 80:   notes.append('⚠ MEM Limit 위험')
        elif mem_pct_req >= 40: notes.append('△ MEM Request 40% 초과')  # Pod 감설 기준
        note = ' / '.join(notes)
        note_bg = C['WARN']  if '⚠' in note else C['WARN2'] if '△' in note else bg
        note_tc = C['RED_T'] if '⚠' in note else C['YEL_T'] if '△' in note else C['SGRAY']

        cells = [
            (nl,          nb,                        nt,                        True,  None,  'center'),
            (str(row.get('cluster','-')), bg,       C['SGRAY'],               False, None,  'left'),
            (str(row.get('namespace','-')),bg,      C['SGRAY'],               False, None,  'left'),
            (str(row.get('rs_or_pod', row.get('podName','-'))),  bg,      C['BLACK'],               False, None,  'left'),
            (str(row.get('onodeName','-')), bg, C['SGRAY'], False, None, 'left'),
            (str(row.get('Deployment','') or '-'), bg, C['SGRAY'], False, None, 'left'),
            (str(row.get('replicaSetName','') or '-'), bg, C['SGRAY'], False, None, 'left'),
            (int(cpu_req), bg,                       C['BLACK'],               False, '0',   'right'),
            (cpu_used,    bg,                        C['BLACK'],               False, '0.0', 'right'),
            (cpu_pct_req, pct_bg(cpu_pct_req) or bg, pct_tc(cpu_pct_req), cpu_pct_req>=40, '0.0', 'right'),
            (cpu_pct_lim, pct_bg(cpu_pct_lim) or bg, pct_tc(cpu_pct_lim), cpu_pct_lim>=40, '0.0', 'right'),
            (mem_req_mb,  bg,                        C['BLACK'],               False, '0.0', 'right'),
            (mem_lim_mb,  bg,                        C['BLACK'],               False, '0.0', 'right'),
            (mem_used_mb, bg,                        C['BLACK'],               False, '0.0', 'right'),
            (mem_pct_req, pct_bg(mem_pct_req) or bg, pct_tc(mem_pct_req), mem_pct_req>=40, '0.0', 'right'),
            (mem_pct_lim, pct_bg(mem_pct_lim) or bg, pct_tc(mem_pct_lim), mem_pct_lim>=40, '0.0', 'right'),
            (note,        note_bg,                   note_tc,               bool(note), None, 'left'),
        ]
        for ci,(v,b,tc,bold,fmt,ha) in enumerate(cells,1):
            wt = ci in (3,4,5,6,7,17)
            sc(ws,R,ci,v,9 if wt else 10,bold,tc,b,ha,wt,fmt)


def _rs_to_deployment(rs_name):
    """
    replicaSetName에서 Deployment명 추출.
    ReplicaSet명 패턴: [deployment]-[rs_hash 9~12자리]
    예) act-was-users-ro-6b7cbf7b85  →  act-was-users-ro
    """
    import re
    m = re.match(r'^(.+)-[a-z0-9]{9,12}$', str(rs_name))
    return m.group(1) if m else rs_name


def _extract_workload_name(pod_name, rs_name=''):
    """
    워크로드명 추출 우선순위:
    1. replicaSetName 있으면 → RS명에서 Deployment 해시 suffix 제거
    2. 없으면 → Pod명 정규식으로 fallback
    """
    import re
    # 1순위: replicaSetName 기반
    rs = str(rs_name).strip()
    if rs and rs not in ('', 'nan', 'None'):
        return _rs_to_deployment(rs)
    # 2순위: Pod명 정규식 (DaemonSet 등 RS 없는 경우)
    m = re.match(r'^(.+)-[a-z0-9]{9,12}-[a-z0-9]{5}$', pod_name)
    if m:
        return m.group(1)
    m2 = re.match(r'^(.+)-[a-z0-9]{5}$', pod_name)
    if m2:
        return m2.group(1)
    return pod_name


def make_pod_summary_sheet(wb, pod_df, date_range=''):
    """
    Pod 요약 시트: ReplicaSet명 기준으로 Pod를 묶어 최대값 표시.
    - CPU Request=0 & 실사용=0  → 구 RS 잔존 → 하단 별도 표
    - CPU Request=0 & 실사용>0  → Request 미설정 → 메인 표, 비고에 표시
    - 그 외                      → 메인 표
    """
    if pod_df.empty:
        return

    import pandas as _pd

    df = pod_df.copy()

    # ── 수치 변환
    NUM_COLS = [
        'cpu_request_val', 'cpu_total_milli_max',
        'cpu_per_request_max', 'cpu_per_limit_max',
        'memory_request_val', 'memory_limit_val', 'mem_working_set_max',
        'memory_per_request_max', 'memory_per_limit_max',
    ]
    for c in NUM_COLS:
        if c in df.columns:
            df[c] = _pd.to_numeric(df[c], errors='coerce').fillna(0)

    CPU_REQ  = 'cpu_request_val'
    CPU_USED = 'cpu_total_milli_max'

    # rs_or_pod: load_pod_files에서 이미 RS 기준으로 집계돼 있음
    # replicaSetName 없는 경우 Pod명 기준으로 집계된 값
    if 'rs_or_pod' not in df.columns:
        df['rs_or_pod'] = df['podName'].apply(lambda p: _extract_workload_name(p, ''))


    # ── 구 RS 판별
    # 기준: active_pod_count / pod_count < 5% (활성 Pod가 거의 없음)
    # 또는 cpu_request_val=0 AND cpu_total_milli_max=0 (완전 비활성)
    ACTIVE_RATIO_THRESHOLD = 0.05  # 5% 미만이면 구 RS로 분류

    if 'active_pod_count' in df.columns and 'pod_count' in df.columns:
        df['_active_ratio'] = _pd.to_numeric(df['active_pod_count'], errors='coerce').fillna(0) / \
                              _pd.to_numeric(df['pod_count'], errors='coerce').replace(0, 1)
        df['_dead'] = (
            ((df[CPU_REQ] == 0) & (df[CPU_USED] == 0)) |   # 완전 비활성
            (df['_active_ratio'] < ACTIVE_RATIO_THRESHOLD)  # 활성 Pod 5% 미만
        )
    else:
        df['_dead'] = (df[CPU_REQ] == 0) & (df[CPU_USED] == 0)

    # Request=0 AND 실사용>0 → Request 미설정 운영 중
    df['_req_unset'] = (df[CPU_REQ] == 0) & (df[CPU_USED] > 0) & ~df['_dead']

    df_active = df[~df['_dead']].copy()
    df_dead   = df[ df['_dead']].copy()

    # ── RS 단위 집계 (namespace + rs_or_pod 기준)
    def do_agg(data):
        agg = {}
        # Pod 수: load_pod_files에서 이미 RS별 pod_count로 집계됨
        if 'pod_count' in data.columns:
            agg['Pod수'] = ('pod_count', 'sum')
        else:
            agg['Pod수'] = ('rs_or_pod', 'count')
        for c in NUM_COLS:
            if c in data.columns: agg[c] = (c, 'max')
        for c in ['netlabel', 'cluster']:
            if c in data.columns: agg[c] = (c, 'first')
        result = (
            data.groupby(['namespace', 'rs_or_pod'])
            .agg(**agg)
            .reset_index()
            .sort_values(['netlabel', 'cluster', 'namespace', 'rs_or_pod'])
            .reset_index(drop=True)
        )
        # _req_unset: Request=0 & 실사용>0 여부를 RS별로 집계
        if '_req_unset' in data.columns:
            ru = data.groupby(['namespace','rs_or_pod'])['_req_unset'].any().reset_index()
            result = result.merge(ru, on=['namespace','rs_or_pod'], how='left')
        return result

    grp = do_agg(df_active)

    # 구 RS 별도 집계
    grp_dead = _pd.DataFrame()
    if not df_dead.empty:
        agg_d = {}
        for c in ['netlabel', 'cluster']:
            if c in df_dead.columns: agg_d[c] = (c, 'first')
        agg_d['Pod수'] = ('pod_count', 'sum') if 'pod_count' in df_dead.columns else ('rs_or_pod', 'count')
        grp_dead = (
            df_dead.groupby(['namespace', 'rs_or_pod'])
            .agg(**agg_d)
            .reset_index()
            .sort_values(['netlabel', 'cluster', 'namespace', 'rs_or_pod'])
            .reset_index(drop=True)
        )

    # ── 시트 생성
    ws = wb.create_sheet('Pod 요약 (RS별)')
    ws.tab_color = 'E65100'
    ws.sheet_view.showGridLines = False

    HDRS = [
        ('망',              7 ),
        ('클러스터',        22),
        ('네임스페이스',    22),
        ('ReplicaSet명',    44),
        ('Pod\n수',          6 ),
        ('CPU\nRequest(m)', 11),
        ('CPU 실사용\n최대(m)', 11),
        ('CPU\nReq 대비(%)', 11),
        ('CPU\nLimit 대비(%)', 11),
        ('MEM\nRequest(MB)', 12),
        ('MEM\nLimit(MB)',   12),
        ('MEM 실사용\n최대(MB)', 12),
        ('MEM\nReq 대비(%)', 11),
        ('MEM\nLimit 대비(%)', 11),
        ('비고', 32),
    ]
    nC = len(HDRS)
    for ci,(h,w) in enumerate(HDRS,1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 제목
    ws.merge_cells(f'A1:{get_column_letter(nC)}1')
    sc(ws,1,1,f'Pod 요약 (ReplicaSet별){" — " + date_range if date_range else ""}',
       13,True,'FFFFFF',C['NAVY'],'left',border=False)
    ws.row_dimensions[1].height = 28

    # 부제
    ws.merge_cells(f'A2:{get_column_letter(nC)}2')
    sc(ws,2,1,
       '※ ReplicaSet 기준 묶음 | CPU/MEM 최대값 | 활성 Pod 5% 미만 또는 Request=0&실사용=0인 RS → 제외→하단 표시 | Request=0&실사용>0은 비고에 ⚠ 표시',
       9,False,'888888',C['NAVY'],'left',border=False)
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 38

    for ci,(h,w) in enumerate(HDRS,1):
        sc(ws,4,ci,h,10,True,'FFFFFF',C['NAVY'],'center',True)
    ws.freeze_panes = 'A5'

    # ── 헬퍼
    def fv(row, key):
        try:
            import pandas as _p
            f = float(row.get(key, 0) or 0)
            return 0.0 if _p.isna(f) else f
        except:
            return 0.0

    def pb(v): return C['WARN'] if v>=80 else C['WARN2'] if v>=40 else None  # Pod: 40%이하 감설
    def pt(v): return C['RED_T'] if v>=80 else C['YEL_T'] if v>=40 else C['BLACK']

    # ── 메인 표 행 출력
    for ri, row in grp.iterrows():
        R = ri + 5
        ws.row_dimensions[R].height = 18
        bg = C['LGRAY'] if ri%2==0 else C['WHITE']
        nl = str(row.get('netlabel',''))
        nb = C['IM_BG'] if nl=='인망' else C['HM_BG']
        nt = C['IM_T']  if nl=='인망' else C['HM_T']

        cpu_req     = fv(row,'cpu_request_val')
        cpu_used    = fv(row,'cpu_total_milli_max')
        cpu_pct_req = fv(row,'cpu_per_request_max')
        cpu_pct_lim = fv(row,'cpu_per_limit_max')
        mem_req_mb  = fv(row,'memory_request_val')
        mem_lim_mb  = fv(row,'memory_limit_val')
        mem_used_mb = fv(row,'mem_working_set_max')
        mem_pct_req = fv(row,'memory_per_request_max')
        mem_pct_lim = fv(row,'memory_per_limit_max')
        pod_cnt     = int(row.get('Pod수',0))
        req_unset   = bool(row.get('_req_unset', False))

        notes = []
        if req_unset:               notes.append('⚠ Request 미설정')
        if cpu_pct_lim >= 80:       notes.append('⚠ CPU Limit 위험')
        elif cpu_pct_req >= 40:     notes.append('△ CPU Req 40% 초과')  # Pod 감설 기준
        if mem_pct_lim >= 80:       notes.append('⚠ MEM Limit 위험')
        elif mem_pct_req >= 40:     notes.append('△ MEM Req 40% 초과')  # Pod 감설 기준
        note    = ' / '.join(notes)
        note_bg = C['WARN']  if '⚠' in note else C['WARN2'] if '△' in note else bg
        note_tc = C['RED_T'] if '⚠' in note else C['YEL_T'] if '△' in note else C['SGRAY']

        cells = [
            (nl,           nb,               nt,              True,  None,  'center'),
            (str(row.get('cluster','-')),   bg, C['SGRAY'],   False, None,  'left'),
            (str(row.get('namespace','-')), bg, C['SGRAY'],   False, None,  'left'),
            (str(row.get('rs_or_pod','-')),      bg, C['BLACK'],   False, None,  'left'),
            (pod_cnt,      bg,               C['SGRAY'],      False, '0',   'center'),
            (int(cpu_req), bg,               C['BLACK'],      False, '0',   'right'),
            (cpu_used,     bg,               C['BLACK'],      False, '0.0', 'right'),
            (cpu_pct_req,  pb(cpu_pct_req) or bg, pt(cpu_pct_req), cpu_pct_req>=40, '0.0', 'right'),
            (cpu_pct_lim,  pb(cpu_pct_lim) or bg, pt(cpu_pct_lim), cpu_pct_lim>=40, '0.0', 'right'),
            (mem_req_mb,   bg,               C['BLACK'],      False, '0.0', 'right'),
            (mem_lim_mb,   bg,               C['BLACK'],      False, '0.0', 'right'),
            (mem_used_mb,  bg,               C['BLACK'],      False, '0.0', 'right'),
            (mem_pct_req,  pb(mem_pct_req) or bg, pt(mem_pct_req), mem_pct_req>=40, '0.0', 'right'),
            (mem_pct_lim,  pb(mem_pct_lim) or bg, pt(mem_pct_lim), mem_pct_lim>=40, '0.0', 'right'),
            (note,         note_bg,          note_tc,         bool(note), None, 'left'),
        ]
        for ci,(v,b,tc,bold,fmt,ha) in enumerate(cells,1):
            wt = ci in (3,4,15)
            sc(ws,R,ci,v,9 if wt else 10,bold,tc,b,ha,wt,fmt)

    # ── 구 RS 잔존 별도 표 (메인 표 아래 3행 공백)
    if not grp_dead.empty:
        base = len(grp) + 5 + 3

        ws.row_dimensions[base].height = 22
        ws.merge_cells(f'A{base}:{get_column_letter(nC)}{base}')
        sc(ws, base, 1,
           f'▼  구 RS 잔존 (활성 Pod 5% 미만 또는 Request=0&실사용=0) — {len(grp_dead)}개 RS / {int(grp_dead["Pod수"].sum())}개 Pod  ▼',
           10, True, 'FFFFFF', 'B45309', 'left', border=False)

        DEAD_HDRS = [('망',7),('클러스터',22),('네임스페이스',22),('ReplicaSet명',44),('Pod\n수',6)]
        ws.row_dimensions[base+1].height = 30
        for ci,(h,w) in enumerate(DEAD_HDRS,1):
            sc(ws, base+1, ci, h, 9, True, 'FFFFFF', 'B45309', 'center', True)

        for ri, row in grp_dead.iterrows():
            R = base + 2 + ri
            ws.row_dimensions[R].height = 16
            bg = C['LGRAY'] if ri%2==0 else C['WHITE']
            nl = str(row.get('netlabel',''))
            nb = C['IM_BG'] if nl=='인망' else C['HM_BG']
            nt = C['IM_T']  if nl=='인망' else C['HM_T']
            cells = [
                (nl, nb, nt, True, None, 'center'),
                (str(row.get('cluster','-')),   bg, C['SGRAY'], False, None, 'left'),
                (str(row.get('namespace','-')), bg, C['SGRAY'], False, None, 'left'),
                (str(row.get('rs_or_pod','-')),      bg, C['BLACK'], False, None, 'left'),
                (int(row.get('Pod수',0)),       bg, C['SGRAY'], False, '0',  'center'),
            ]
            for ci,(v,b,tc,bold,fmt,ha) in enumerate(cells,1):
                sc(ws,R,ci,v,9,bold,tc,b,ha,ci in (3,4),fmt)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('data_dir',  nargs='?', default='.')
    parser.add_argument('spec_file', nargs='?', default=None)
    parser.add_argument('--cpu-target', type=float, default=70)
    parser.add_argument('--mem-target', type=float, default=70)
    parser.add_argument('--out', default=None)
    args = parser.parse_args()
    TC = args.cpu_target; TM = args.mem_target

    print(f"\n📊 인스턴스 리소스 분석  (CPU 목표 {TC}% / MEM 목표 {TM}%)")

    data_path = Path(args.data_dir)
    files = list(data_path.glob('*.xlsx'))

    # 사양 파일 자동 감지
    spec_path = args.spec_file
    if not spec_path:
        for f in files:
            if is_spec_file(f.name):
                spec_path = str(f); print(f"  💡 사양 파일 자동 감지: {f.name}"); break
    if spec_path and not Path(spec_path).exists() and (data_path/spec_path).exists():
        spec_path = str(data_path/spec_path)

    print(f"\n📋 사양 파일: {spec_path or '없음'}")
    spec = load_spec(spec_path)
    has_spec = len(spec) > 0

    # 데이터 분류 및 로드
    print(f"\n🔄 데이터 로드 중...")
    buckets = {('인망','cpu','max'):[],'인망_cpu_avg':[],'인망_net_max':[],'인망_net_avg':[],
               '행망_cpu_max':[],'행망_cpu_avg':[],'행망_net_max':[],'행망_net_avg':[]}
    # 딕셔너리 재정의
    bkt = {}
    pod_files = []
    for f in files:
        # Pod 파일 우선 감지 (podName 컬럼 존재 여부로 판단)
        if is_pod_file(f):
            pod_files.append(f)
            print(f"  📦 Pod: {f.name[:50]}")
            continue
        cls = classify_file(f.name)
        if cls:
            key = f'{cls[0]}_{cls[1]}_{cls[2]}'
            bkt.setdefault(key,[]).append(f)
            print(f"  ✓ {f.name[:50]}")

    def load_merge(flist, vcols):
        if not flist: return pd.DataFrame()
        dfs = []
        for f in flist:
            try: dfs.append(load_xlsx(str(f), vcols))
            except Exception as e: print(f"  ❌ {f.name}: {e}")
        if not dfs: return pd.DataFrame()
        df = pd.concat(dfs, ignore_index=True)
        return df.drop_duplicates(subset=['oname','time'])

    cm_im = load_merge(bkt.get('인망_cpu_max',[]), ['cpu','memory_pused'])
    ca_im = load_merge(bkt.get('인망_cpu_avg',[]), ['cpu','memory_pused'])
    cm_hm = load_merge(bkt.get('행망_cpu_max',[]), ['cpu','memory_pused'])
    ca_hm = load_merge(bkt.get('행망_cpu_avg',[]), ['cpu','memory_pused'])

    def agg_cpu(mdf, adf, netlabel):
        if mdf.empty: return pd.DataFrame()
        m = mdf.groupby('oname').agg(cpu_max=('cpu','max'), mem_max=('memory_pused','max')).reset_index()
        a = adf.groupby('oname').agg(cpu_avg=('cpu','mean'), mem_avg=('memory_pused','mean')).reset_index() if not adf.empty else pd.DataFrame()
        df = m.merge(a, on='oname', how='left') if not a.empty else m.assign(cpu_avg=0, mem_avg=0)
        df['netlabel'] = netlabel
        df[['cpu_max','cpu_avg','mem_max','mem_avg']] = df[['cpu_max','cpu_avg','mem_max','mem_avg']].fillna(0).round(1)
        return df

    im = agg_cpu(cm_im, ca_im, '인망')
    hm = agg_cpu(cm_hm, ca_hm, '행망')
    df = pd.concat([im, hm], ignore_index=True)
    print(f"\n✅ 집계: {len(df)}대 (인망 {len(im)}, 행망 {len(hm)})")

    # 80% 발생 날짜
    all_max = pd.concat([cm_im, cm_hm], ignore_index=True)
    if not all_max.empty:
        daily = all_max.groupby(['oname','time']).agg(cd=('cpu','max'),md=('memory_pused','max')).reset_index()
        def alert_col(col, alias):
            return daily[daily[col]>=80].groupby('oname')['time'].apply(
                lambda x: ', '.join(sorted(str(d) for d in x))).reset_index().rename(columns={'time':alias})
        df = df.merge(alert_col('cd','cpu_alert'), on='oname', how='left')
        df = df.merge(alert_col('md','mem_alert'), on='oname', how='left')
    df['cpu_alert'] = df.get('cpu_alert', pd.Series(['-']*len(df))).fillna('-')
    df['mem_alert'] = df.get('mem_alert', pd.Series(['-']*len(df))).fillna('-')

    # 스펙 매핑
    if has_spec:
        def get_spec_val(r, key):
            s = lookup_spec(spec, r['netlabel'], r['oname'])
            return s.get(key) if s else None
        df['시스템명']  = df.apply(lambda r: get_spec_val(r,'시스템명'), axis=1)
        df['현재_cpu'] = df.apply(lambda r: get_spec_val(r,'cpu'), axis=1)
        df['현재_ram'] = df.apply(lambda r: get_spec_val(r,'ram'), axis=1)
        print(f"  스펙 매핑: {df['현재_cpu'].notna().sum()}/{len(df)}대")

    # 서버유형 분류
    def _server_type(oname):
        o = str(oname).lower()
        if any(k in o for k in ['clst-worker','cluster-worker','ap-clst','fe-dmz']) and 'cicd' not in o: return '워커노드'
        if 'gateway' in o and 'worker' in o: return '게이트웨이'
        if 'kubectl' in o: return 'K8s관리'
        if 'redis' in o: return 'Redis'
        if 'tibero' in o: return '물리DB'
        if 'cicd' in o: return 'CICD워커'
        return '일반서버'
    df['서버유형'] = df['oname'].apply(_server_type)

    # 판정 (list comprehension으로 pandas tuple 언팩 방지)
    TH = {'s_max':40, 'r_max':40, 'keep':70}  # VM 기준: 40%이하 감설 / 70%이상 증설
    _grades = [grade_row(row, DEFAULT_RULES, TH) for _, row in df.iterrows()]
    df['판정']   = [g[0] for g in _grades]
    df['reason'] = [g[1] for g in _grades]

    # 권고 계산
    if has_spec:
        _recs = [calc_rec(row, TC, TM) for _, row in df.iterrows()]
        df['권고_cpu']   = [r[0] for r in _recs]
        df['권고_ram']   = [r[1] for r in _recs]
        df['스케일방향'] = [r[2] for r in _recs]
        df['예상_cpu']   = [r[3] for r in _recs]
        df['예상_mem']   = [r[4] for r in _recs]
    else:
        df['권고_cpu']   = None; df['권고_ram']   = None
        df['스케일방향'] = None; df['예상_cpu']   = None; df['예상_mem'] = None

    print("\n📈 판정 결과:")
    for k,v in df['판정'].value_counts().items(): print(f"  {k}: {v}대")
    if has_spec and '스케일방향' in df.columns:
        print("\n📉 스케일 방향:")
        for k,v in df['스케일방향'].value_counts().items(): print(f"  {k}: {v}대")

    # 날짜 범위
    dates = all_max['time'].dropna().tolist() if not all_max.empty else []
    dr = f"{min(dates)} ~ {max(dates)}" if dates else "기간 미확인"

    # ── 엑셀 생성
    print(f"\n📝 엑셀 보고서 생성 중...")
    wb = Workbook()
    ws1 = wb.active; ws1.title = '요약 대시보드'
    ws1.sheet_view.showGridLines = False

    ws1.column_dimensions['A'].width = 14
    for i,w in enumerate([14,16,12,16,14,14,12,10],2): ws1.column_dimensions[get_column_letter(i)].width=w

    ws1.merge_cells('A1:I1')
    sc(ws1,1,1,f'인스턴스 리소스 사용률 분석  ({dr})',14,True,'FFFFFF',C['NAVY'],'left',border=False)
    ws1.row_dimensions[1].height=30
    ws1.merge_cells('A2:I2')
    sc(ws1,2,1,f'최한시·주말 제외 | 최대값 기준 판정 | 목표: CPU {TC}% / MEM {TM}% | NHN Cloud vCPU: 2→4→8→16→32→64',
       10,False,'888888',C['NAVY'],'left',border=False)
    ws1.row_dimensions[2].height=16

    pj_keys = ['강력권고','검토','일시적피크→검토','유지(주의)','유지(확정)','예외처리']
    pj_hdrs = ['구분','🟢 강력권고','🟡 검토','🟠 일시피크→검토','🔴 유지(주의)','🔵 유지(확정)','⚪ 예외처리','전체']
    ws1.row_dimensions[4].height=32
    for ci,h in enumerate(pj_hdrs,1): sc(ws1,4,ci,h,10,True,'FFFFFF',C['BLUE'],'center',True)
    for ri,(nl,nb,nt) in enumerate([('인망',C['IM_BG'],C['IM_T']),('행망',C['HM_BG'],C['HM_T'])],5):
        ws1.row_dimensions[ri].height=22
        sc(ws1,ri,1,nl,11,True,nt,nb)
        for ci,pj in enumerate(pj_keys,2):
            cnt = len(df[(df['netlabel']==nl)&(df['판정']==pj)])
            pb,pt = PJ_FILL.get(pj,(C['WHITE'],C['BLACK']))
            sc(ws1,ri,ci,cnt,12,cnt>0,pt,pb)
        sc(ws1,ri,8,len(df[df['netlabel']==nl]),12,True,C['BLACK'],C['LGRAY'])

    # 정렬: 인망→행망, 각 hostname 알파벳 순
    df['_no'] = df['netlabel'].map({'인망':0,'행망':1}).fillna(2)
    df_sorted = df.sort_values(['_no','oname']).drop(columns='_no').reset_index(drop=True)

    # 투트랙: 간소 + 상세
    make_detail_sheet(wb, df_sorted, '전체 상세(간소)', '1B5E9E', mode='simple', tc=TC, tm=TM)
    make_detail_sheet(wb, df_sorted, '전체 상세(상세)', '2E4D8C', mode='full',   tc=TC, tm=TM)

    # 스케일 다운 후보
    cand = df_sorted[df_sorted['스케일방향'].str.startswith('스케일 다운', na=False)]
    make_detail_sheet(wb, cand, f'스케일 다운 ({len(cand)}대)', '00AA44', mode='full', tc=TC, tm=TM)

    # 스케일 업 후보 (유지(주의) -> 스케일 업)
    up_cand = df_sorted[df_sorted['스케일방향'] == '스케일 업']
    if len(up_cand) > 0:
        make_detail_sheet(wb, up_cand, f'스케일 업 ({len(up_cand)}대)', 'C62828', mode='full', tc=TC, tm=TM)

    # 현상태유지(다운불가)
    hold = df_sorted[df_sorted['스케일방향']=='현상태유지(다운불가)']
    if len(hold) > 0:
        make_detail_sheet(wb, hold, f'다운불가-현상태유지 ({len(hold)}대)', 'FF6600', mode='full', tc=TC, tm=TM)

    exc = df_sorted[df_sorted['판정']=='예외처리']
    make_detail_sheet(wb, exc, f'예외처리 ({len(exc)}대)', '9E9E9E', mode='simple', tc=TC, tm=TM)

    # ── Pod 분석 시트
    if pod_files:
        print(f"\n📦 Pod 데이터 집계 중... ({len(pod_files)}개 파일)")
        pod_df, _pod_dates = load_pod_files(pod_files)
        pod_dr = ", ".join(d.strftime("%Y-%m-%d") for d in _pod_dates) if _pod_dates else dr
        make_pod_sheet(wb, pod_df, pod_dr)
        make_pod_summary_sheet(wb, pod_df, pod_dr)
        print(f"  ✓ Pod 분석 시트 생성 완료 (상세 + 워크로드 요약)")
    else:
        print(f"\n📦 Pod 파일 없음 — Pod 분석 시트 생략 (파일 업로드 시 자동 생성)")

    out_name = args.out or f'리소스_분석보고서_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    wb.save(out_name)
    print(f"\n✅ 저장 완료: {out_name}")
    print(f"   시트: {', '.join(s.title for s in wb.worksheets)}")

if __name__ == '__main__':
    main()
