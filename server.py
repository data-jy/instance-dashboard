#!/usr/bin/env python3
"""
리소스 분석 웹 서버
사용법: python server.py [--port 8000]
브라우저: http://localhost:8000
"""
import sys, os, json, tempfile, threading, webbrowser, argparse
import unicodedata
from pathlib import Path
from http.server import HTTPServer, BaseHTTPRequestHandler
from io import BytesIO
from urllib.parse import parse_qs, urlparse

# analyze.py의 핵심 함수들 import
sys.path.insert(0, str(Path(__file__).parent))
try:
    from analyze import (
        nfc, classify_file, is_spec_file, is_pod_file, load_pod_files,
        load_xlsx, load_spec, lookup_spec,
        DEFAULT_RULES, grade_row, calc_rec,
        make_sheet, make_detail_sheet, make_pod_sheet, make_pod_summary_sheet,
        NHN_CPU, NHN_SPECS, find_best_spec,
        _rs_to_deployment, _extract_workload_name,
        HOLD_BG, HOLD_T, GROUP_NOW, GROUP_REC
    )
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"❌ analyze.py 또는 패키지 오류: {e}")
    print("   analyze.py와 같은 폴더에 server.py를 두고 실행하세요.")
    sys.exit(1)

# ─────────────────────────────────────────
# HTML UI (인라인 — index.html 없어도 동작)
# ─────────────────────────────────────────
HTML_UI = None  # 아래에서 로드

def load_html():
    global HTML_UI
    # 1) 같은 폴더의 resource_analyze.html 우선
    candidates = [
        Path(__file__).parent / 'resource_analyze.html',
        Path(__file__).parent / 'index.html',
        Path(__file__).parent / 'static' / 'index.html',
    ]
    for p in candidates:
        if p.exists():
            HTML_UI = p.read_text(encoding='utf-8')
            print(f"  ✓ UI 파일 로드: {p.name}")
            return
    # 2) 없으면 미니멀 내장 UI
    HTML_UI = BUILTIN_HTML
    print("  💡 내장 미니멀 UI 사용 (resource_analyze.html을 같은 폴더에 두면 기존 UI 사용)")

# ─────────────────────────────────────────
# 핵심: 분석 실행 함수
# ─────────────────────────────────────────
def run_analysis(data_files: dict, spec_bytes: bytes, spec_name: str,
                 config: dict) -> bytes:
    """
    data_files: { 'filename.xlsx': bytes, ... }
    spec_bytes: 사양 파일 bytes (없으면 None)
    config: { cpu_target, mem_target, nic_im, nic_hm, rules: [...] }
    반환: xlsx bytes
    """
    TC = float(config.get('cpu_target', 70))
    TM = float(config.get('mem_target', 70))
    NIC_IM = float(config.get('nic_im', 10)) * 1024
    NIC_HM = float(config.get('nic_hm', 2))  * 1024
    custom_rules = config.get('rules', [])  # [{kw, action, reason}, ...]

    # 서비스 현황 (빈 문자열이면 '-' 처리)
    svc_raw = config.get('svc', {})
    def _sv(key): return svc_raw.get(key, '') or '-'
    svc = {
        'im': [_sv('im_avg_user'), _sv('im_max_user'), _sv('im_con_avg'), _sv('im_con_max'), _sv('im_avg_tps'), _sv('im_max_tps')],
        'hm': [_sv('hm_avg_user'), _sv('hm_max_user'), _sv('hm_con_avg'), _sv('hm_con_max'), _sv('hm_avg_tps'), _sv('hm_max_tps')],
    }

    # 규칙: 커스텀 우선, 없으면 DEFAULT
    rules = [(r['kw'], r['action'], r['reason']) for r in custom_rules if r.get('kw')]

    # ── 사양 파일 파싱
    spec = {}
    has_spec = False
    if spec_bytes:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(spec_bytes)
            tmp = f.name
        try:
            spec = load_spec(tmp)
            has_spec = len(spec) > 0
        finally:
            os.unlink(tmp)

    # ── 데이터 파일 분류 및 로드
    bkt = {}
    pod_tmp_files = []   # Pod 임시 파일 경로
    tmp_files = []
    print(f"  📂 수신 파일 목록:")
    for fname in data_files.keys():
        cls = classify_file(fname)
        print(f"    [{fname!r}] → {cls}")
    for fname, fbytes in data_files.items():
        # Pod 파일인지 먼저 확인 (임시 저장 후 podName 컬럼 여부로 판단)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            f.write(fbytes)
            tmp_path = f.name
        if is_pod_file(tmp_path):
            pod_tmp_files.append(tmp_path)
            tmp_files.append(tmp_path)
            print(f"  📦 Pod 파일 감지: {fname}")
            continue
        cls = classify_file(fname)
        if not cls:
            os.unlink(tmp_path)  # 분류 불가 파일 즉시 제거
            continue
        key = f'{cls[0]}_{cls[1]}_{cls[2]}'
        tmp_files.append(tmp_path)
        bkt.setdefault(key, []).append(tmp_path)

    def load_merge(flist, vcols):
        if not flist: return pd.DataFrame()
        dfs = []
        for p in flist:
            try: dfs.append(load_xlsx(p, vcols))
            except Exception as e: print(f"  ❌ {p}: {e}")
        if not dfs: return pd.DataFrame()
        df = pd.concat(dfs, ignore_index=True)
        return df.drop_duplicates(subset=['oname','time'])

    cm_im = load_merge(bkt.get('인망_cpu_max',[]), ['cpu','memory_pused'])
    ca_im = load_merge(bkt.get('인망_cpu_avg',[]), ['cpu','memory_pused'])
    cm_hm = load_merge(bkt.get('행망_cpu_max',[]), ['cpu','memory_pused'])
    ca_hm = load_merge(bkt.get('행망_cpu_avg',[]), ['cpu','memory_pused'])
    nm_im = load_merge(bkt.get('인망_net_max',[]), ['trafficIn','trafficOut'])
    na_im = load_merge(bkt.get('인망_net_avg',[]), ['trafficIn','trafficOut'])
    nm_hm = load_merge(bkt.get('행망_net_max',[]), ['trafficIn','trafficOut'])
    na_hm = load_merge(bkt.get('행망_net_avg',[]), ['trafficIn','trafficOut'])

    # 임시 파일 정리 (Pod 파일은 나중에 쓰므로 제외)
    for p in tmp_files:
        if p in pod_tmp_files:
            continue  # Pod 파일은 아직 삭제하지 않음
        try: os.unlink(p)
        except: pass

    def agg_cpu(mdf, adf, netlabel):
        if mdf.empty: return pd.DataFrame()
        m = mdf.groupby('oname').agg(cpu_max=('cpu','max'), mem_max=('memory_pused','max')).reset_index()
        a = adf.groupby('oname').agg(cpu_avg=('cpu','mean'), mem_avg=('memory_pused','mean')).reset_index() if not adf.empty else pd.DataFrame()
        df2 = m.merge(a, on='oname', how='left') if not a.empty else m.assign(cpu_avg=0, mem_avg=0)
        df2['netlabel'] = netlabel
        df2[['cpu_max','cpu_avg','mem_max','mem_avg']] = df2[['cpu_max','cpu_avg','mem_max','mem_avg']].fillna(0).round(1)
        return df2

    def agg_net(mdf, adf, netlabel, nic_mbps):
        if mdf.empty: return pd.DataFrame()
        m = mdf.groupby('oname').agg(in_max=('trafficIn','max'), out_max=('trafficOut','max')).reset_index()
        a = adf.groupby('oname').agg(in_avg=('trafficIn','mean'), out_avg=('trafficOut','mean')).reset_index() if not adf.empty else pd.DataFrame()
        df2 = m.merge(a, on='oname', how='left') if not a.empty else m.assign(in_avg=0, out_avg=0)
        df2['netlabel'] = netlabel
        for c in ['in_max','out_max','in_avg','out_avg']:
            df2[c] = (df2[c].fillna(0) / 1e6).round(1)
        df2['in_util_pct']  = (df2['in_max']  / nic_mbps * 100).round(1)
        df2['out_util_pct'] = (df2['out_max'] / nic_mbps * 100).round(1)
        return df2

    im = agg_cpu(cm_im, ca_im, '인망')
    hm = agg_cpu(cm_hm, ca_hm, '행망')
    df = pd.concat([im, hm], ignore_index=True)

    if df.empty:
        raise ValueError("분석 가능한 CPU/MEM 데이터가 없습니다. 파일명 형식을 확인하세요. (예: YYMMDD-YYMMDD_인망_운영_cpu_mem_최대값.xlsx)")
    im_net = agg_net(nm_im, na_im, '인망', NIC_IM)
    hm_net = agg_net(nm_hm, na_hm, '행망', NIC_HM)

    # 80% 발생 날짜
    all_max = pd.concat([cm_im, cm_hm], ignore_index=True)
    if not all_max.empty:
        daily = all_max.groupby(['oname','time']).agg(cd=('cpu','max'), md=('memory_pused','max')).reset_index()
        def alert_col(col, alias):
            return daily[daily[col]>=80].groupby('oname')['time'].apply(
                lambda x: ', '.join(sorted(str(d) for d in x))).reset_index().rename(columns={'time':alias})
        df = df.merge(alert_col('cd','cpu_alert'), on='oname', how='left')
        df = df.merge(alert_col('md','mem_alert'), on='oname', how='left')
    df['cpu_alert'] = df.get('cpu_alert', pd.Series(['-']*len(df))).fillna('-')
    df['mem_alert'] = df.get('mem_alert', pd.Series(['-']*len(df))).fillna('-')

    # 스펙 매핑
    if has_spec:
        df['시스템명']  = df.apply(lambda r: (lookup_spec(spec,r['netlabel'],r['oname']) or {}).get('시스템명'), axis=1)
        df['현재_cpu'] = df.apply(lambda r: (lookup_spec(spec,r['netlabel'],r['oname']) or {}).get('cpu'), axis=1)
        df['현재_ram'] = df.apply(lambda r: (lookup_spec(spec,r['netlabel'],r['oname']) or {}).get('ram'), axis=1)
    else:
        df['시스템명'] = None; df['현재_cpu'] = None; df['현재_ram'] = None

    # 서버유형 분류
    def _server_type(oname):
        o = str(oname).lower()
        if any(k in o for k in ['clst-worker','cluster-worker','ap-clst','fe-dmz']) and 'cicd' not in o: return '워커노드'
        if 'gateway' in o and 'worker' in o: return '게이트웨이'
        if 'kubectl' in o: return 'K8s관리'
        if 'redis' in o: return 'Redis'
        if 'tibero' in o: return '물리DB'
        if 'cicd' in o: return 'CICD워커'
        return '일반서버'
    df['서버유형'] = df['oname'].apply(_server_type)

    # 판정 (pandas가 tuple을 DataFrame으로 언팩하는 것을 방지하기 위해 list로 감싸서 처리)
    TH = {'s_max':40, 'r_max':40, 'keep':70}  # VM 기준: 40%이하 감설 / 70%이상 증설
    _grades = [grade_row(row, rules, TH) for _, row in df.iterrows()]
    df['판정']   = [g[0] for g in _grades]
    df['reason'] = [g[1] for g in _grades]

    print(f"  📊 판정 분포: {df['판정'].value_counts().to_dict()}")
    print(f"  📊 전체 서버 수: 인망 {len(df[df['netlabel']=='인망'])}대 / 행망 {len(df[df['netlabel']=='행망'])}대")

    # 권고 계산
    if has_spec:
        _recs = [calc_rec(row, TC, TM) for _, row in df.iterrows()]
        df['권고_cpu']   = [r[0] for r in _recs]
        df['권고_ram']   = [r[1] for r in _recs]
        df['스케일방향'] = [r[2] for r in _recs]
        df['예상_cpu']   = [r[3] for r in _recs]
        df['예상_mem']   = [r[4] for r in _recs]
    else:
        df['권고_cpu']   = None; df['권고_ram']   = None
        df['스케일방향'] = None; df['예상_cpu']   = None; df['예상_mem'] = None

    # 날짜 범위
    dates = all_max['time'].dropna().tolist() if not all_max.empty else []
    dr = f"{min(dates)} ~ {max(dates)}" if dates else "기간 미확인"

    # 엑셀 생성
    wb = Workbook()
    ws1 = wb.active; ws1.title = '요약 대시보드'
    ws1.sheet_view.showGridLines = False

    from openpyxl.styles import Font as F2, PatternFill as PF2, Alignment as A2, Border as B2, Side as S2
    thin = S2(style='thin', color='CCCCCC')
    def bdr(): return B2(left=thin,right=thin,top=thin,bottom=thin)
    def fl(c): return PF2('solid', start_color=c, end_color=c)
    C2 = dict(NAVY='1F3864',BLUE='2E4D8C',GRN='C8E6C9',GRN_T='166534',
              YEL='FFD54F',YEL_T='854D0E',RED='FF8A80',RED_T='991B1B',
              TEAL='B2EBF2',TEAL_T='155E75',ORG='FFE0B2',ORG_T='9A3412',
              LGRAY='F5F5F5',WHITE='FFFFFF',BLACK='1A1F33',SGRAY='5A6480',
              IM_BG='DBEAFE',IM_T='1E40AF',HM_BG='DCFCE7',HM_T='166534',
              DN_BG='E3F2FD',DN_T='1565C0',UP_BG='FCE4EC',UP_T='B71C1C',
              WARN='FFD0D0',WARN2='FFF8D0')

    def sc2(ws2, r, c, val, sz=10, bold=False, color='1A1F33', bg=None,
            ha='center', wrap=False, fmt=None):
        addr = get_column_letter(c)+str(r)
        ws2[addr] = val
        ws2[addr].font = F2(name='Arial',sz=sz,bold=bold,color=color)
        if bg: ws2[addr].fill = fl(bg)
        ws2[addr].alignment = A2(horizontal=ha,vertical='center',wrap_text=wrap)
        ws2[addr].border = bdr()
        if fmt: ws2[addr].number_format = fmt

    PJ_FILL2 = {
        '강력권고':(C2['GRN'],C2['GRN_T']), '검토':(C2['YEL'],C2['YEL_T']),
        '유지(확정)':(C2['TEAL'],C2['TEAL_T']), '유지(주의)':(C2['RED'],C2['RED_T']),
        '일시적피크→검토':(C2['ORG'],C2['ORG_T']),
        '예외처리':(C2['LGRAY'],C2['SGRAY']), '보통':(C2['WHITE'],C2['BLACK']),
    }

    # ── 요약 대시보드 생성
    for i,w in enumerate([14,16,12,16,14,8,12,10],1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    def _sec_title(r, title):
        ws1.merge_cells(f'A{r}:H{r}')
        ws1[f'A{r}'] = title
        ws1[f'A{r}'].font = F2(name='Arial',sz=12,bold=True,color='FFFFFF')
        ws1[f'A{r}'].fill = fl(C2['NAVY'])
        ws1[f'A{r}'].alignment = A2(horizontal='left',vertical='center')
        ws1.row_dimensions[r].height = 24

    # 제목
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f'인스턴스 리소스 사용률 분석  ({dr})'
    ws1['A1'].font = F2(name='Arial',sz=14,bold=True,color=C2['NAVY'])
    ws1['A1'].alignment = A2(horizontal='left',vertical='center')
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells('A2:H2')
    ws1['A2'] = f'최한시·주말 제외 | 최대값 기준 | 목표: CPU {TC}% / MEM {TM}% | NHN Cloud vCPU: 2→4→8→16→32→64'
    ws1['A2'].font = F2(name='Arial',sz=10,color='888888')
    ws1['A2'].alignment = A2(horizontal='left',vertical='center')
    ws1.row_dimensions[2].height = 16
    ws1.row_dimensions[3].height = 8

    R = 4  # 현재 행 포인터

    # ── 서비스 운영 현황 (값이 있을 때만 표시)
    has_svc = any(v != '-' for v in svc['im'] + svc['hm'])
    if has_svc:
        _sec_title(R, '서비스 운영 현황'); R += 1
        svc_hdrs = ['구분','이용자 평균','이용자 최대','동시접속 평균','동시접속 최대','처리량 평균','처리량 최대']
        ws1.row_dimensions[R].height = 28
        for ci,h in enumerate(svc_hdrs,1):
            sc2(ws1,R,ci,h,10,True,'FFFFFF',C2['BLUE'],'center',True)
        R += 1
        for (nl,nb,nt,vals) in [('인터넷망',C2['IM_BG'],C2['IM_T'],svc['im']),
                                  ('행정망',  C2['HM_BG'],C2['HM_T'],svc['hm'])]:
            ws1.row_dimensions[R].height = 22
            sc2(ws1,R,1,nl,11,True,nt,nb)
            for ci,v in enumerate(vals,2):
                sc2(ws1,R,ci,v,10,False,C2['BLACK'],nb)
            R += 1
        ws1.row_dimensions[R].height = 10; R += 1

    # ── 판정 현황 테이블
    _sec_title(R, '최종 판정 현황 (서버 역할 반영)'); R += 1
    pj_keys = ['강력권고','검토','일시적피크→검토','유지(주의)','유지(확정)','예외처리']
    pj_hdrs = ['구분','🟢 강력권고','🟡 검토','🟠 일시피크→검토','🔴 유지(주의)','🔵 유지(확정)','⚪ 예외처리','전체']
    ws1.row_dimensions[R].height = 32
    for ci,h in enumerate(pj_hdrs,1):
        sc2(ws1,R,ci,h,10,True,'FFFFFF',C2['BLUE'],'center',True)
    R += 1
    for (nl,nb,nt) in [('인망',C2['IM_BG'],C2['IM_T']),('행망',C2['HM_BG'],C2['HM_T'])]:
        ws1.row_dimensions[R].height = 22
        sc2(ws1,R,1,nl,11,True,nt,nb)
        for ci2,pj in enumerate(pj_keys,2):
            cnt = len(df[(df['netlabel']==nl)&(df['판정']==pj)])
            pb,pt = PJ_FILL2.get(pj,(C2['WHITE'],C2['BLACK']))
            sc2(ws1,R,ci2,cnt,12,cnt>0,pt,pb)
        sc2(ws1,R,8,len(df[df['netlabel']==nl]),12,True,C2['BLACK'],C2['LGRAY'])
        R += 1
    ws1.row_dimensions[R].height = 10; R += 1

    # ── 스케일 방향 테이블
    _sec_title(R, '스케일 방향 현황'); R += 1
    sd_hdrs = ['구분','🔽 스케일 다운','🔼 스케일 업','⏸ 현상태유지(다운불가)','▶ 유지(확정/주의)','⚪ 예외처리/기타','전체',None]
    ws1.row_dimensions[R].height = 32
    for ci,h in enumerate(sd_hdrs,1):
        if h: sc2(ws1,R,ci,h,10,True,'FFFFFF',C2['BLUE'],'center',True)
    R += 1
    for (nl,nb,nt) in [('인망',C2['IM_BG'],C2['IM_T']),('행망',C2['HM_BG'],C2['HM_T'])]:
        ws1.row_dimensions[R].height = 22
        sc2(ws1,R,1,nl,11,True,nt,nb)
        dn  = len(df[(df['netlabel']==nl)&(df['스케일방향'].str.startswith('스케일 다운', na=False))])
        up  = len(df[(df['netlabel']==nl)&(df['스케일방향']=='스케일 업')])
        hld = len(df[(df['netlabel']==nl)&(df['스케일방향']=='현상태유지(다운불가)')])
        keep= len(df[(df['netlabel']==nl)&(df['판정'].isin(['유지(확정)','보통']))])
        tot = len(df[df['netlabel']==nl])
        etc = tot - dn - up - hld - keep
        sc2(ws1,R,2,dn, 12,dn>0, C2['DN_T'],C2['DN_BG'])
        sc2(ws1,R,3,up, 12,up>0, C2['RED_T'],C2['RED'])
        sc2(ws1,R,4,hld,12,hld>0,C2['ORG_T'],C2['WARN'])
        sc2(ws1,R,5,keep,12,keep>0,C2['TEAL_T'],C2['TEAL'])
        sc2(ws1,R,6,etc,12,False,C2['SGRAY'],C2['LGRAY'])
        sc2(ws1,R,7,tot,12,True, C2['BLACK'],C2['LGRAY'])
        R += 1
    ws1.row_dimensions[R].height = 10; R += 1

    # ── 판정 기준 테이블
    _sec_title(R, '판정 기준'); R += 1
    sc2(ws1,R,1,'판정',10,True,'FFFFFF',C2['BLUE'])
    ws1.merge_cells(f'B{R}:H{R}')
    sc2(ws1,R,2,'설명',10,True,'FFFFFF',C2['BLUE'],'left')
    ws1.row_dimensions[R].height = 20; R += 1

    legend = [
        ('🟢 강력권고',    C2['GRN'],  C2['GRN_T'],  'CPU 최대 < 50% AND MEM 최대 < 60% (최대값 기준 판정)'),
        ('🟡 검토',        C2['YEL'],  C2['YEL_T'],  'CPU 최대 < 60% AND MEM 최대 < 70%'),
        ('🟠 일시피크→검토',C2['ORG'],  C2['ORG_T'],  '순간 피크 패턴 — 평균은 낮으나 배포/배치 시 최대 급등. 스케일 다운 검토 가능'),
        ('🔴 유지(주의)',  C2['RED'],  C2['RED_T'],  'CPU 또는 MEM 최대 ≥ 70% — 증설 검토 대상 (스케일 업 권고)'),
        ('🔵 유지(확정)',  C2['TEAL'], C2['TEAL_T'], '서버 역할 확인 후 유지 결정 (ionengs, static, tracer, pdf/edms, whatap, ahnlab, cicd 등)'),
        ('⚪ 예외처리',    C2['LGRAY'],C2['SGRAY'],  'WAF, 방화벽, IPS, 테스트서버, 신규 서버(수집 미흡) 등'),
    ]
    for (pj,pb,pt,desc) in legend:
        ws1.row_dimensions[R].height = 20
        sc2(ws1,R,1,pj,10,True,pt,pb)
        ws1.merge_cells(f'B{R}:H{R}')
        sc2(ws1,R,2,desc,10,False,C2['BLACK'],C2['WHITE'],'left')
        R += 1

    # 정렬: 인망→행망, 각 hostname 알파벳
    df['_no'] = df['netlabel'].map({'인망':0,'행망':1}).fillna(2)
    df_sorted = df.sort_values(['_no','oname']).drop(columns='_no').reset_index(drop=True)

    make_detail_sheet(wb, df_sorted, '전체 상세(간소)', '1B5E9E', mode='simple', tc=TC, tm=TM)
    make_detail_sheet(wb, df_sorted, '전체 상세(상세)', '2E4D8C', mode='full',   tc=TC, tm=TM)

    cand = df_sorted[df_sorted['스케일방향'].str.startswith('스케일 다운', na=False)]
    make_detail_sheet(wb, cand, f'스케일 다운 ({len(cand)}대)', '00AA44', mode='full', tc=TC, tm=TM)

    # 스케일 업 후보 (유지(주의) -> 스케일 업)
    up_cand = df_sorted[df_sorted['스케일방향'] == '스케일 업']
    if len(up_cand) > 0:
        make_detail_sheet(wb, up_cand, f'스케일 업 ({len(up_cand)}대)', 'C62828', mode='full', tc=TC, tm=TM)

    hold = df_sorted[df_sorted['스케일방향']=='현상태유지(다운불가)']
    if len(hold) > 0:
        make_detail_sheet(wb, hold, f'다운불가-현상태유지 ({len(hold)}대)', 'FF6600', mode='full', tc=TC, tm=TM)

    # 2~8코어 집중 분석
    if has_spec:
        def _safe_cpu(x):
            try: return int(float(x))
            except: return 0
        small = df_sorted[
            df_sorted['현재_cpu'].apply(_safe_cpu).isin([2,4,8]) &
            ~df_sorted['판정'].isin(['예외처리'])
        ]
        if len(small) > 0:
            make_detail_sheet(wb, small, f'2~8코어 집중 분석 ({len(small)}대)', 'FF6600', mode='full', tc=TC, tm=TM)

    # 워커노드 분석
    workers = df_sorted[df_sorted['서버유형']=='워커노드']
    if len(workers) > 0:
        make_detail_sheet(wb, workers, f'워커노드 분석 ({len(workers)}대)', '5B2C6F', mode='full', tc=TC, tm=TM)

    # Pod 분석 시트
    if pod_tmp_files:
        print(f"  \n  📦 Pod 데이터 집계 중... ({len(pod_tmp_files)}개 파일)")
        pod_df, _pod_dates = load_pod_files(pod_tmp_files)
        pod_dr = ", ".join(d.strftime("%Y-%m-%d") for d in _pod_dates) if _pod_dates else dr
        make_pod_sheet(wb, pod_df, pod_dr)
        make_pod_summary_sheet(wb, pod_df, pod_dr)
        # Pod 임시 파일 정리
        for p in pod_tmp_files:
            try: os.unlink(p)
            except: pass
    

    exc = df_sorted[df_sorted['판정']=='예외처리']
    make_detail_sheet(wb, exc, f'예외처리 ({len(exc)}대)', '9E9E9E', mode='simple', tc=TC, tm=TM)

    # 메모리에 저장
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────
# 내장 미니멀 UI (resource_analyze.html 없을 때 fallback)
# ─────────────────────────────────────────
BUILTIN_HTML = """<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8">
<title>리소스 분석</title>
<style>
body{font-family:'Noto Sans KR',sans-serif;background:#F0F4FF;margin:0;padding:40px}
.card{background:#fff;border-radius:14px;padding:32px;max-width:720px;margin:0 auto;box-shadow:0 2px 16px rgba(0,0,0,.08)}
h1{color:#1F3864;font-size:22px;margin:0 0 8px}
.sub{color:#888;font-size:13px;margin-bottom:28px}
.section{margin-bottom:20px}
.section label{display:block;font-weight:600;font-size:13px;color:#374151;margin-bottom:8px}
.drop-zone{border:2px dashed #93C5FD;border-radius:10px;padding:24px;text-align:center;
  background:#EFF6FF;cursor:pointer;transition:.15s;color:#1E40AF;font-size:14px}
.drop-zone:hover,.drop-zone.drag{background:#DBEAFE;border-color:#3B82F6}
.drop-zone.has-files{border-color:#4ADE80;background:#F0FDF4;color:#166534}
.spec-zone{border:2px dashed #86EFAC;border-radius:10px;padding:16px;text-align:center;
  background:#F0FDF4;cursor:pointer;transition:.15s;color:#166534;font-size:13px}
.spec-zone:hover{background:#DCFCE7}
.spec-zone.loaded{border-color:#4ADE80;font-weight:600}
.config{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.config input{width:100%;padding:8px 12px;border:1.5px solid #D1D5DB;border-radius:8px;
  font-size:14px;box-sizing:border-box;outline:none}
.config input:focus{border-color:#3B82F6}
.btn{width:100%;padding:14px;background:#1F3864;color:#fff;border:none;border-radius:10px;
  font-size:15px;font-weight:700;cursor:pointer;margin-top:8px;transition:.15s}
.btn:hover:not(:disabled){background:#2E4D8C}
.btn:disabled{opacity:.5;cursor:not-allowed}
.progress{display:none;margin-top:16px;padding:14px;background:#F8FAFF;border-radius:8px;
  font-size:13px;color:#374151;border:1px solid #E5E7EB}
.prog-bar-wrap{background:#E5E7EB;border-radius:4px;height:6px;margin:8px 0}
.prog-bar{background:#3B82F6;height:6px;border-radius:4px;transition:width .3s;width:0%}
.file-list{margin-top:10px;font-size:12px;color:#4B5563}
.tag{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600;margin:2px}
</style></head><body>
<div class="card">
  <h1>📊 인스턴스 리소스 분석</h1>
  <div class="sub">와탭 데이터 + 사양 파일을 업로드하면 엑셀 보고서를 생성합니다</div>

  <div class="section">
    <label>① 인스턴스 사양 파일 <span style="font-weight:400;color:#9CA3AF">(선택 — 시스템명·권고 사양 자동 계산)</span></label>
    <div class="spec-zone" id="specZone" onclick="document.getElementById('specInput').click()"
      ondragover="ev.preventDefault();this.classList.add('drag')"
      ondragleave="this.classList.remove('drag')"
      ondrop="ev.preventDefault();this.classList.remove('drag');setSpecFile(ev.dataTransfer.files[0])">
      <span id="specLabel">📋 클릭하거나 드래그 — 인스턴스 사양 xlsx</span>
    </div>
    <input type="file" id="specInput" accept=".xlsx" style="display:none"
      onchange="setSpecFile(this.files[0]);this.value=''">
  </div>

  <div class="section">
    <label>② 와탭 데이터 파일 <span style="font-weight:400;color:#9CA3AF">(필수 — 기간당 8개, 여러 기간 동시 가능)</span></label>
    <div class="drop-zone" id="dropZone" onclick="document.getElementById('dataInput').click()"
      ondragover="event.preventDefault();this.classList.add('drag')"
      ondragleave="this.classList.remove('drag')"
      ondrop="event.preventDefault();this.classList.remove('drag');addFiles(event.dataTransfer.files)">
      📁 클릭하거나 드래그하여 xlsx 업로드<br>
      <span style="font-size:12px;opacity:.7">인망/행망 × CPU-MEM/Network × 최대/평균값</span>
    </div>
    <input type="file" id="dataInput" multiple accept=".xlsx" style="display:none"
      onchange="addFiles(this.files);this.value=''">
    <div class="file-list" id="fileList"></div>
  </div>

  <div class="section">
    <label>③ 분석 설정</label>
    <div class="config">
      <div><div style="font-size:12px;color:#6B7280;margin-bottom:4px">CPU 목표 사용률 (%)</div>
        <input type="number" id="cpuTarget" value="70" min="50" max="90"></div>
      <div><div style="font-size:12px;color:#6B7280;margin-bottom:4px">MEM 목표 사용률 (%)</div>
        <input type="number" id="memTarget" value="70" min="50" max="90"></div>
      <div><div style="font-size:12px;color:#6B7280;margin-bottom:4px">인터넷망 NIC (Gbps)</div>
        <input type="number" id="nicIM" value="10"></div>
      <div><div style="font-size:12px;color:#6B7280;margin-bottom:4px">행정망 NIC (Gbps)</div>
        <input type="number" id="nicHM" value="2"></div>
    </div>
  </div>

  <button class="btn" id="runBtn" onclick="runAnalysis()" disabled>⚡ 분석 실행 → 엑셀 다운로드</button>

  <div class="progress" id="progress">
    <div id="progMsg">분석 중...</div>
    <div class="prog-bar-wrap"><div class="prog-bar" id="progBar"></div></div>
  </div>
</div>

<script>
let dataFiles = [];
let specFile = null;

function setSpecFile(f) {
  if (!f || !f.name.endsWith('.xlsx')) return;
  specFile = f;
  const z = document.getElementById('specZone');
  z.classList.add('loaded');
  document.getElementById('specLabel').textContent = '✅ ' + f.name;
}

function addFiles(flist) {
  [...flist].forEach(f => {
    if (!f.name.endsWith('.xlsx')) return;
    if (!dataFiles.find(x => x.name === f.name)) dataFiles.push(f);
  });
  renderFiles();
}

const TYPE_LABELS = {
  'cpu_max':'CPU/MEM MAX','cpu_avg':'CPU/MEM AVG',
  'net_max':'NET MAX','net_avg':'NET AVG'
};
function guessType(name) {
  const n = name.toLowerCase();
  const t = n.includes('cpu_mem') ? 'cpu' : n.includes('network') ? 'net' : null;
  const s = n.includes('최대') ? 'max' : n.includes('평균') ? 'avg' : null;
  return t && s ? TYPE_LABELS[t+'_'+s] || '?' : '분류불가';
}

function renderFiles() {
  const el = document.getElementById('fileList');
  const btn = document.getElementById('runBtn');
  if (!dataFiles.length) { el.innerHTML=''; btn.disabled=true; return; }
  btn.disabled = false;
  const z = document.getElementById('dropZone');
  z.classList.toggle('has-files', dataFiles.length > 0);
  z.innerHTML = `📁 ${dataFiles.length}개 파일 업로드됨 (클릭으로 추가)<br><span style="font-size:12px;opacity:.7">파일명을 클릭하면 제거</span>`;
  el.innerHTML = dataFiles.map((f,i) =>
    `<span class="tag" style="background:#DBEAFE;color:#1E40AF;cursor:pointer" onclick="removeFile(${i})" title="클릭하여 제거">
      ${f.name.length>40 ? f.name.slice(0,37)+'...' : f.name} <b>${guessType(f.name)}</b> ✕
    </span>`
  ).join('');
}

function removeFile(i) { dataFiles.splice(i,1); renderFiles(); }

async function runAnalysis() {
  if (!dataFiles.length) return;
  const btn = document.getElementById('runBtn');
  const prog = document.getElementById('progress');
  const msg  = document.getElementById('progMsg');
  const bar  = document.getElementById('progBar');
  btn.disabled=true; prog.style.display='block';

  const form = new FormData();
  dataFiles.forEach(f => form.append('data_files', f));
  if (specFile) form.append('spec_file', specFile);
  form.append('config', JSON.stringify({
    cpu_target: +document.getElementById('cpuTarget').value,
    mem_target: +document.getElementById('memTarget').value,
    nic_im:     +document.getElementById('nicIM').value,
    nic_hm:     +document.getElementById('nicHM').value,
    rules: []
  }));

  msg.textContent = '⏳ 분석 중... (보통 3~5초)'; bar.style.width='20%';
  let animInterval = setInterval(() => {
    const cur = parseFloat(bar.style.width);
    if (cur < 85) bar.style.width = (cur + 2) + '%';
  }, 200);

  try {
    const res = await fetch('/analyze', { method:'POST', body:form });
    clearInterval(animInterval);
    if (!res.ok) {
      const err = await res.json();
      msg.textContent = '❌ 오류: ' + (err.detail || err.error || '알 수 없는 오류');
      bar.style.width = '0%';
    } else {
      bar.style.width = '100%';
      msg.textContent = '✅ 완료! 다운로드 중...';
      const blob = await res.blob();
      const url = URL.createObjectURL(blob);
      const a = document.createElement('a');
      a.href = url;
      const today = new Date();
      a.download = '리소스_분석보고서_' + today.getFullYear() +
        String(today.getMonth()+1).padStart(2,'0') +
        String(today.getDate()).padStart(2,'0') + '.xlsx';
      a.click();
      URL.revokeObjectURL(url);
    }
  } catch(e) {
    clearInterval(animInterval);
    msg.textContent = '❌ 서버 연결 오류: ' + e.message;
  }
  btn.disabled = false;
}
</script></body></html>"""


# ─────────────────────────────────────────
# HTTP 핸들러
# ─────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        print(f"  [{self.address_string()}] {fmt % args}")

    def do_GET(self):
        parsed = urlparse(self.path)
        # ping — 서버 연결 확인용
        if parsed.path == '/ping':
            body = b'{"status":"ok"}'
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)
            return
        if parsed.path in ('/', '/index.html'):
            body = HTML_UI.encode('utf-8')
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)
        else:
            self.send_error(404)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', '*')
        self.end_headers()

    def do_POST(self):
        if self.path != '/analyze':
            self.send_error(404); return

        # multipart 파싱
        import cgi
        ctype = self.headers.get('Content-Type','')
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length)

        # cgi.FieldStorage 사용
        import io
        environ = {
            'REQUEST_METHOD': 'POST',
            'CONTENT_TYPE': ctype,
            'CONTENT_LENGTH': str(length),
        }
        fp = io.BytesIO(body)
        form = cgi.FieldStorage(fp=fp, environ=environ)

        try:
            # 데이터 파일들
            data_files = {}
            if 'data_files' in form:
                items = form['data_files']
                if not isinstance(items, list):
                    items = [items]
                for item in items:
                    if hasattr(item, 'filename') and item.filename:
                        data_files[item.filename] = item.file.read()

            # 사양 파일
            spec_item = form['spec_file'] if 'spec_file' in form else None
            spec_bytes = spec_item.file.read() if spec_item is not None and hasattr(spec_item, 'filename') and spec_item.filename else None
            spec_name  = spec_item.filename if spec_item is not None and hasattr(spec_item, 'filename') and spec_item.filename else None

            # 설정
            config_raw = form.getvalue('config', '{}')
            config = json.loads(config_raw)

            print(f"\n  📊 분석 시작: 데이터 {len(data_files)}개, 사양 {'있음' if spec_bytes else '없음'}")

            xlsx_bytes = run_analysis(data_files, spec_bytes, spec_name, config)

            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename*=UTF-8\'\'%EB%A6%AC%EC%86%8C%EC%8A%A4_%EB%B6%84%EC%84%9D%EB%B3%B4%EA%B3%A0%EC%84%9C.xlsx')
            self.send_header('Content-Length', len(xlsx_bytes))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(xlsx_bytes)
            print(f"  ✅ 완료 ({len(xlsx_bytes)//1024}KB)")

        except Exception as e:
            import traceback
            traceback.print_exc()
            error_body = json.dumps({'error': str(e)}).encode()
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Content-Length', len(error_body))
            self.end_headers()
            self.wfile.write(error_body)


# ─────────────────────────────────────────
# 서버 실행
# ─────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description='리소스 분석 웹 서버')
    parser.add_argument('--port', type=int, default=8000, help='포트 번호 (기본 8000)')
    parser.add_argument('--no-browser', action='store_true', help='브라우저 자동 열기 비활성화')
    args = parser.parse_args()

    load_html()

    server = HTTPServer(('0.0.0.0', args.port), Handler)
    url = f"http://localhost:{args.port}"

    print(f"""
╔══════════════════════════════════════════╗
║     리소스 분석 서버 실행 중             ║
╠══════════════════════════════════════════╣
║  브라우저 접속: {url:<27}║
║  종료: Ctrl+C                           ║
╚══════════════════════════════════════════╝

  📌 같은 폴더에 resource_analyze.html 을 두면
     기존 UI를 그대로 사용합니다.
""")

    if not args.no_browser:
        threading.Timer(1.0, lambda: webbrowser.open(url)).start()

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n\n서버 종료.")

if __name__ == '__main__':
    main()
